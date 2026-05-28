#!/usr/bin/env python3
"""
excel_export.py — Excel 评分卡导出（含雷达图）

职责：读取 score.py 输出的 JSON 结果，生成 Excel 评分卡。
依赖：openpyxl
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import RadarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip3 install openpyxl")
    sys.exit(1)


# ============================================================
# 样式定义
# ============================================================

HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="003366")
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, color="666666")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
SCORE_FONT = Font(name="Microsoft YaHei", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# 分数颜色
SCORE_COLORS = {
    "high": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),   # >=8 绿
    "mid": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),     # 6-8 黄
    "low": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),     # <6 红
    "na": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),      # N/A 灰
}

GRADE_COLORS = {
    "A+": "1B5E20", "A": "2E7D32", "B+": "F9A825", "B": "FF8F00",
    "C": "E65100", "F": "B71C1C",
}


def _score_fill(score) -> PatternFill:
    """根据分数返回背景色"""
    if score is None:
        return SCORE_COLORS["na"]
    try:
        s = float(score)
    except (ValueError, TypeError):
        return SCORE_COLORS["na"]
    if s >= 8:
        return SCORE_COLORS["high"]
    elif s >= 6:
        return SCORE_COLORS["mid"]
    else:
        return SCORE_COLORS["low"]


# ============================================================
# 导出函数
# ============================================================

def export_excel(result_json_path: str, output_path: str = None) -> str:
    """
    导出 Excel 评分卡。

    Args:
        result_json_path: score.py 输出的 JSON 文件路径
        output_path: 输出 Excel 路径（默认同目录）

    Returns:
        输出文件路径
    """
    with open(result_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    meta = data["meta"]
    dimensions = data["dimensions"]
    ds = data.get("data_source_assessment", {})
    suggestions = data.get("optimization_suggestions", [])

    if output_path is None:
        base = Path(result_json_path).stem
        output_dir = Path(result_json_path).parent
        output_path = str(output_dir / f"{base}_评分卡.xlsx")

    wb = Workbook()

    # ---- Sheet 1: 评分总览 ----
    ws1 = wb.active
    ws1.title = "评分总览"

    # 标题
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"📊 报告评分卡 — {meta['report_name']}"
    ws1["A1"].font = TITLE_FONT

    ws1.merge_cells("A2:F2")
    ws1["A2"] = (f"模板: {meta.get('template', 'default')} | "
                 f"模型: {meta.get('model', 'N/A')} | "
                 f"日期: {meta.get('score_date', 'N/A')}")
    ws1["A2"].font = SUBTITLE_FONT

    # 总分
    ws1.merge_cells("A4:B4")
    ws1["A4"] = "总分"
    ws1["A4"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    ws1["C4"] = f"{meta['total_score']}/10"
    ws1["C4"].font = Font(name="Microsoft YaHei", size=14, bold=True,
                          color=GRADE_COLORS.get(meta["grade"], "333333"))
    ws1["D4"] = f"{meta['grade']} — {meta['grade_label']}"
    ws1["D4"].font = Font(name="Microsoft YaHei", size=12,
                          color=GRADE_COLORS.get(meta["grade"], "333333"))

    # 维度评分表
    headers = ["维度", "分数", "权重", "加权分", "评价", "改进建议"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=6, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 加载权重
    from score import load_config
    try:
        config = load_config(meta.get("template", "default"))
        weight_map = {d["id"]: d.get("weight", 1.0) for d in config["dimensions"]}
    except Exception:
        weight_map = {}

    for i, dim in enumerate(dimensions):
        row = 7 + i
        score = dim.get("score")
        weight = weight_map.get(dim["id"], 1.0)

        try:
            s = float(score) if score is not None else None
        except (ValueError, TypeError):
            s = None

        weighted = round(s * weight, 2) if s is not None else "N/A"

        values = [
            dim.get("name", dim.get("id", "")),
            s if s is not None else "N/A",
            weight,
            weighted,
            dim.get("evaluation", ""),
            dim.get("improvement", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.alignment = Alignment(horizontal="center")

        # 分数单元格着色
        ws1.cell(row=row, column=2).fill = _score_fill(score)
        ws1.cell(row=row, column=2).font = SCORE_FONT

    # 列宽
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 40
    ws1.column_dimensions["F"].width = 40

    # 雷达图
    chart = RadarChart()
    chart.type = "filled"
    chart.title = "维度评分雷达图"
    chart.style = 26

    labels = Reference(ws1, min_col=1, min_row=7, max_row=6 + len(dimensions))
    values_ref = Reference(ws1, min_col=2, min_row=6, max_row=6 + len(dimensions))
    chart.add_data(values_ref, titles_from_data=True)
    chart.set_categories(labels)

    chart.width = 18
    chart.height = 14
    ws1.add_chart(chart, f"A{9 + len(dimensions)}")

    # ---- Sheet 2: 数据源评估 ----
    ws2 = wb.create_sheet("数据源评估")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "📡 数据来源质量评估"
    ws2["A1"].font = TITLE_FONT

    ds_headers = ["统计项", "数量/分数"]
    for col, h in enumerate(ds_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    ds_rows = [
        ("发现引用总数", ds.get("sources_found", 0)),
        ("权威来源", ds.get("classified", {}).get("authoritative", 0)),
        ("可靠来源", ds.get("classified", {}).get("reliable", 0)),
        ("可疑来源", ds.get("classified", {}).get("questionable", 0)),
        ("不可验证", ds.get("classified", {}).get("unverifiable", 0)),
        ("数据质量分", f"{ds.get('quality_score', 0)}/10"),
    ]
    for i, (label, val) in enumerate(ds_rows):
        ws2.cell(row=4 + i, column=1, value=label).font = NORMAL_FONT
        ws2.cell(row=4 + i, column=2, value=val).font = NORMAL_FONT
        ws2.cell(row=4 + i, column=1).border = THIN_BORDER
        ws2.cell(row=4 + i, column=2).border = THIN_BORDER

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    # ---- Sheet 3: 优化建议 ----
    ws3 = wb.create_sheet("优化建议")

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "🔧 优化建议"
    ws3["A1"].font = TITLE_FONT

    sugg_headers = ["优先级", "领域", "当前状态", "改进方案", "预期影响"]
    for col, h in enumerate(sugg_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    priority_colors = {
        "critical": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "high": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    }

    for i, s in enumerate(suggestions):
        if not s:
            continue
        row = 4 + i
        values = [
            s.get("priority", ""),
            s.get("area", ""),
            s.get("current_state", ""),
            s.get("suggestion", ""),
            s.get("impact", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)

        # 优先级着色
        p = s.get("priority", "")
        if p in priority_colors:
            ws3.cell(row=row, column=1).fill = priority_colors[p]

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 40
    ws3.column_dimensions["E"].width = 25

    # 保存
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="导出 Excel 评分卡")
    parser.add_argument("json_file", help="评分结果 JSON 文件")
    parser.add_argument("--output", help="输出路径")
    args = parser.parse_args()

    path = export_excel(args.json_file, args.output)
    print(f"✅ Excel 评分卡已保存: {path}")
